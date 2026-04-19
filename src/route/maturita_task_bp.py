from flask import Blueprint, request, send_file
from src.models.Maturita import Maturita
from src.models.Maturita_Task import Maturita_Task
from src.models.Task import Task
from src.models.User_Team import User_Team
from src.models.User import User
from src.models.Topic import Topic
from src.models.Team import Team
import flask_login
import datetime
from src.utils.response import send_response
from app import max_INT, limiter
from src.utils.enums import Status, Role
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from src.models.Evaluator import Evaluator
from src.utils.excel import make_borders, center
import io
from src.utils.redis_cache import get_cache, set_cache

#TODO: přidat export té tabulky, nejlíp excel tabulka

maturita_task_bp = Blueprint("maturita_task", __name__)

@maturita_task_bp.route("/maturita_task/get/table", methods = ["GET"])
@flask_login.login_required
@limiter.limit("30/minute")
def get_table():
    amountForPaging = request.args.get("amountForPaging", None)
    pageNumber = request.args.get("pageNumber", None)

    allTasks = []

    now = datetime.datetime.now(tz = datetime.timezone.utc)
    maturita = Maturita.query.filter(Maturita.startDate <= now, now <= Maturita.endDate).first()

    if not maturita:
        return send_response(400, 83010, {"message":"There is no current maturita"}, "error")
    if not amountForPaging:
        return send_response(400, 83020, {"message": "amountForPaging not entered"}, "error")

    try:
        amountForPaging = int(amountForPaging)
    except:
        return send_response(400, 83030, {"message": "amountForPaging not integer"}, "error")
    
    if amountForPaging < 1:
        return send_response(400, 83040, {"message": "amountForPaging smaller than 1"}, "error")
    
    if amountForPaging > max_INT:
        return send_response(400, 83050, {"message": "amountForPaging too big"}, "error")
    
    if not pageNumber:
        return send_response(400, 83060, {"message": "pageNumber not entered"}, "error")
    
    try:
        pageNumber = int(pageNumber)
    except:
        return send_response(400, 83070, {"message": "pageNumber not integer"}, "error")
    if pageNumber > max_INT + 1:
        return send_response(400, 83080, {"message": "pageNumber too big"}, "error")
    
    pageNumber -= 1

    if pageNumber < 0:
        return send_response(400, 83090, {"message": "pageNumber must be bigger than 0"}, "error")

    cacheKey = f"maturita_task:table:{amountForPaging}:{pageNumber}"
    cacheData = get_cache(cacheKey)

    if not cacheData:
        maturitaTask = Maturita_Task.query.join(Team, (Team.guarantor == Maturita_Task.guarantor) & (Team.idTask == Maturita_Task.idTask) & (Team.status == Status.Approved)).filter(Maturita_Task.idMaturita == maturita.id).offset(amountForPaging * pageNumber).limit(amountForPaging)
        count = Maturita_Task.query.filter(Maturita_Task.idMaturita == maturita.id).count()

        for taskMaturita in maturitaTask:
            task = Task.query.filter_by(id = taskMaturita.idTask, guarantor = taskMaturita.guarantor).first()
            guarantor = User.query.filter_by(id = taskMaturita.guarantor).first()
            userTeam = User_Team.query.filter_by(idTask = taskMaturita.idTask, guarantor = taskMaturita.guarantor).first()
            student = User.query.filter_by(id = userTeam.idUser).first()
            objector = User.query.filter_by(id = taskMaturita.objector).first()
            topic = Topic.query.filter_by(id = taskMaturita.idTopic).first()

            if not objector:
                objectorData = {}
            else:
                objectorData = {
                    "id":objector.id,
                    "name":objector.name,
                    "surname": objector.surname,
                    "abbreviation": objector.abbreviation,
                    "createdAt": objector.createdAt,
                    "role": objector.role.value,
                    "profilePicture":objector.profilePicture,
                    "email":objector.email
                    }

            guarantorData = {
                "id":guarantor.id,
                "name":guarantor.name,
                "surname": guarantor.surname,
                "abbreviation": guarantor.abbreviation,
                "createdAt": guarantor.createdAt,
                "role": guarantor.role.value,
                "profilePicture":guarantor.profilePicture,
                "email":guarantor.email
                }

            allTasks.append({
                "user":{
                    "id":student.id,
                    "name":student.name,
                    "surname": student.surname,
                    "abbreviation": student.abbreviation,
                    "createdAt": student.createdAt,
                    "role": student.role.value,
                    "profilePicture":student.profilePicture,
                    "email":student.email
                },
                "topic":{
                    "id":topic.id,
                    "name":topic.name,
                },
                "variant":taskMaturita.variant,
                "task":task.name,
                "guarantor":guarantorData,
                "objector":objectorData
            })
        set_cache(cacheKey, {"tasks":allTasks, "count":count})

    else:
        allTasks = cacheData["tasks"]
        count = cacheData["count"]

    return send_response(200, 83101, {"message":"All tasks for current maturita", "tasks":allTasks, "count":count}, "success")

@maturita_task_bp.route("/maturita_task/get/excel", methods = ["GET"])
@flask_login.login_required
@limiter.limit("3/minute")
def get_excel():
    if flask_login.current_user.role == Role.Student:
        return send_response(403, 109010, {"message": "No permission for that"}, "error")

    now = datetime.datetime.now(tz = datetime.timezone.utc)
    maturita = Maturita.query.filter(Maturita.startDate <= now, now <= Maturita.endDate).first()

    maturitaTask = Maturita_Task.query.join(Team, (Team.guarantor == Maturita_Task.guarantor) & (Team.idTask == Maturita_Task.idTask) & (Team.status == Status.Approved)).filter(Maturita_Task.idMaturita == maturita.id).all()
    
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "zaci"
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 30
    ws1.column_dimensions["F"].width = 10
    ws1.column_dimensions["G"].width = 40
    ws1.column_dimensions["H"].width = 20
    ws1.column_dimensions["I"].width = 20

    ws1.append(["Id", "Příjmení", "Jméno", "Téma\n(číslo)", "Téma\n(název)", "Varianta", "Varianta (slovně)", "Vedoucí", "Oponent"])

    center(ws1, 9)

    for taskMaturita in maturitaTask:
        task = Task.query.filter_by(id = taskMaturita.idTask, guarantor = taskMaturita.guarantor).first()
        guarantor = User.query.filter_by(id = taskMaturita.guarantor).first()
        userTeam = User_Team.query.filter_by(idTask = taskMaturita.idTask, guarantor = taskMaturita.guarantor).first()
        student = User.query.filter_by(id = userTeam.idUser).first()
        objector = User.query.filter_by(id = taskMaturita.objector).first()
        topic = Topic.query.filter_by(id = taskMaturita.idTopic).first()

        if not objector:
            objectorAbbreviation = None
        else:
            objectorAbbreviation = objector.abbreviation


        ws1.append([student.id, student.surname, student.name, topic.id, topic.name, taskMaturita.variant, task.name, guarantor.abbreviation, objectorAbbreviation])

    ws1.auto_filter.ref = ws1.dimensions

    make_borders(ws1, border)

    evaluators = Evaluator.query.filter_by(idMaturita = maturita.id).order_by(Evaluator.idUser.asc()).all()

    ws2 = wb.create_sheet(title = "ucitele")
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 50
    ws2.column_dimensions["C"].width = 20

    ws2.append(["Id", "Učitel", "Zkratka"])

    center(ws2, 3)

    for evaluator in evaluators:
        user = User.query.filter_by(id = evaluator.idUser).first()
        ws2.append([user.id, user.surname + " " + user.name, user.abbreviation])

    make_borders(ws2, border)

    topics = Topic.query.order_by(Topic.id.asc()).all()

    ws3 = wb.create_sheet(title = "temata")
    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 50

    ws3.append(["Id", "Název"])

    center(ws3, 2)

    for topic in topics:
        ws3.append([topic.id, topic.name])

    make_borders(ws3, border)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name=f"{now.strftime('%y%m%d')}_seznam_zaku.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")