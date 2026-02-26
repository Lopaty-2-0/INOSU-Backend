from openpyxl.styles import Alignment, Font

def make_borders(ws, border):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.font = Font(color=None) 

def center(ws, columns):
    for i in range(columns):
        ws[f"{chr(65 + i)}1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)